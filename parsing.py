"""見積書PDF/Excelからの品目（品名・数量・単価）抽出ロジック。

PDFはClaudeにファイルを直接読ませてJSON形式で品目を取得する。
Claude Code CLI（claude コマンド）を subprocess で呼び出す。
Anthropic APIキーは不要。Claude Pro/Maxサブスクリプションのみで動作する。

Excel（.xlsx）はopenpyxlでヘッダー列を検出して読み込む。
"""
from __future__ import annotations

import io
import json
import re
import subprocess
import tempfile
from pathlib import Path

from openpyxl import load_workbook

# ── 共通ユーティリティ ──────────────────────────────────────────────

ITEM_NAME_KEYWORDS = ["品名", "品目", "名称", "商品名", "item", "description"]
QTY_KEYWORDS = ["数量", "個数", "qty", "quantity"]
UNIT_KEYWORDS = ["単位", "unit"]
SPEC_KEYWORDS = ["規格", "仕様", "寸法", "spec"]
PRICE_KEYWORDS = ["単価", "価格", "unit price", "price"]
AMOUNT_KEYWORDS = ["金額", "amount", "total"]

NUM_RE = re.compile(r"[¥￥]?\s*(-?[0-9][0-9,]*(?:\.[0-9]+)?)\s*円?")

_SUMMARY_ROW_MARKERS = ("小計", "合計", "中計", "大計", "【", "】")


def _normalize_text(s: str) -> str:
    return re.sub(r"[\s　]+", "", s).lower()


def to_number(value) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    if not s:
        return None
    m = NUM_RE.search(s)
    if not m:
        return None
    try:
        return float(m.group(1).replace(",", ""))
    except ValueError:
        return None


def find_col(header_texts: list[str], keywords: list[str]) -> int | None:
    norm_headers = [_normalize_text(h) for h in header_texts]
    norm_keywords = [_normalize_text(k) for k in keywords]
    for i, h in enumerate(norm_headers):
        for kw in norm_keywords:
            if kw and kw in h:
                return i
    return None


# ── PDF: Claude Code CLI による品目抽出 ──────────────────────────────

CLAUDE_BIN = "claude"
CLAUDE_TIMEOUT_SEC = 1800  # 30分。API混雑時も対応できる範囲で設定

_CLAUDE_PROMPT = """\
日本語の見積書PDFです（ファイル名: {filename}）。
レイアウトはPDFごとに異なります。内容を読み取り、JSONのみを返してください。

出力形式（このJSONのみ・説明文不要）:
{{
  "大項目": "見積書全体を一括りにする単一の工事名（例: 給湯室改修工事）。＜＞括弧は含めない。なければ空文字",
  "items": [{{"no":"","種別":"","品名":"","規格仕様":"","備考":"","数量":"","単位":"","単価":"","金額":""}}],
  "notes": "【別途見積】や【見積条件】など品目表の外にある特記事項・注意書きのテキスト（改行は\\nで表現）"
}}

大項目（トップレベルフィールド）の注意事項:
・書類全体が1つの工事名で括れる場合のみ入れる（例: 書類タイトルや表頭に「○○工事」と書かれている場合）
・品目表の中に ＜クロス貼替工事＞ ＜床工事＞ のような複数の ＜＞ 見出し行がある場合は、
  このフィールドは空文字にし、各 ＜＞ 行を後述のとおり items 内に種別="大項目" として含める

items の注意事項:
・no には、品目表の「No.」列にある番号を文字列でそのまま入れる（例: "no": "4"）
・No.列が空の行（付属品・サブ品目・仕様補足など、番号が振られていない行）は "no": "" とし、
  直前の番号付き品目と同じ種別にする。独立した品目として扱わない
・＜クロス貼替工事＞ ＜床工事＞ などの ＜＞ 囲みの見出し行は品目として items に含め、
  種別を "大項目"、品名に括弧を除いた見出し名を入れる。数量・単価・金額はすべて空文字
  例: {{"no":"","種別":"大項目","品名":"クロス貼替工事","規格仕様":"","備考":"","数量":"","単位":"","単価":"","金額":""}}
・＜＞ 見出しの直下にある番号付き品目（例: 1. 玄関 クロス貼替 壁）は no に番号、種別は空文字でよい
・種別には、大項目の下の中分類カテゴリ名を入れる（例: 「業務用厨房機器 ﾏﾙﾅﾝ工業」「水栓 KVK」）。
  ＜＞ 見出し直下の品目で中分類がない場合は種別は空文字でよい
・グループ見出し（例: 「業務用厨房機器 ﾏﾙﾅﾝ工業」）の下に並ぶ品目は、品番や仕入れ先が異なっていても
  同じ種別でまとめる。グループ区切り（罫線・空行・次の見出し）が来るまでは種別を変えない
  （例: 「業務用厨房機器」見出し直下の湯沸かし器・コンロ・配管類もすべて同じ種別にする）
・同じグループ内の複数品目は種別を共通にし、品名にはグループ名を繰り返さない
  （例: 「業務用厨房機器 ﾏﾙﾅﾝ工業」グループ内の「ｺﾝﾛｷｬﾋﾞﾈｯﾄ」→ 品名は「ｺﾝﾛｷｬﾋﾞﾈｯﾄ」のみ、種別に「業務用厨房機器 ﾏﾙﾅﾝ工業」）
・中分類グループのどれにも属さない単体の品目だけ種別を空文字にする
・小計・合計・中計・大計・消費税など集計行は items に含めない
・値引きがあれば品名を「値引」とし金額をマイナス値で含める
・数量・単価・金額は数字のみ（カンマや円マークは不要）

notes の注意事項:
・【別途見積】【見積条件】など、品目表の外に記載されている文章をそのまま入れる
・特記事項が無ければ空文字でよい
"""


class ClaudeExtractionError(RuntimeError):
    pass


def _run_claude(cmd: list[str], timeout: int, cwd: str | None = None) -> dict:
    """Claude CLI を実行して結果 dict を返す。"""
    try:
        proc = subprocess.run(
            cmd, cwd=cwd, capture_output=True, text=True, timeout=timeout
        )
    except FileNotFoundError as e:
        raise ClaudeExtractionError(
            "`claude` コマンドが見つかりません。Claude Code CLI がインストールされ、"
            "PATH が通っていることを確認してください。"
        ) from e
    except subprocess.TimeoutExpired as e:
        raise ClaudeExtractionError(
            f"Claude による解析が{timeout}秒を超えたため中断しました。"
            "APIが混雑している可能性があります。しばらく待ってから再試行してください。"
        ) from e

    if proc.returncode != 0:
        raise ClaudeExtractionError(
            f"Claude の呼び出しに失敗しました（終了コード {proc.returncode}）。"
            f"\n{proc.stderr.strip()[:500]}"
        )
    try:
        return json.loads(proc.stdout)
    except json.JSONDecodeError as e:
        raise ClaudeExtractionError(
            "Claude の応答をJSONとして解釈できませんでした。"
        ) from e


def _recover_partial_json(json_str: str, raw_text: str) -> dict:
    """途中で切れたJSON文字列から品目オブジェクトを可能な限り復元する。

    Claude のレスポンスが長すぎて途中で打ち切られた場合に呼ばれる。
    完結している品目オブジェクト（{"no":..., "品名":...} 形式）を正規表現で
    抽出し、items リストとして返す。大項目・notes は空文字扱いとする。
    """
    # 大項目・notes は先頭近くにあることが多いので先に拾う
    daikomi_m = re.search(r'"大項目"\s*:\s*"([^"]*)"', json_str)
    daikomi = daikomi_m.group(1) if daikomi_m else ""
    notes_m = re.search(r'"notes"\s*:\s*"((?:[^"\\]|\\.)*)"', json_str)
    notes = notes_m.group(1).replace("\\n", "\n") if notes_m else ""

    # 品目オブジェクト: "品名" キーを含む {} ブロックを貪欲でなく抽出
    item_pattern = re.compile(r'\{[^{}]*?"品名"\s*:\s*"[^{}]*?\}', re.DOTALL)
    items = []
    for m in item_pattern.finditer(json_str):
        try:
            obj = json.loads(m.group(0))
            if isinstance(obj, dict) and obj.get("品名"):
                items.append(obj)
        except json.JSONDecodeError:
            pass

    if not items:
        raise ClaudeExtractionError(
            f"Claude の応答をJSONとして解釈できませんでした。\n"
            f"応答先頭: {raw_text[:300]}"
        )

    return {"大項目": daikomi, "items": items, "notes": notes}


def _parse_claude_result(result: dict) -> tuple[list[dict], str, str, float | None]:
    """Claude の result dict から品目リスト・備考テキスト・大項目名・コストを取り出す。"""
    if result.get("is_error"):
        raise ClaudeExtractionError(
            f"Claude がエラーを返しました: {result.get('result')}"
        )
    raw_text = result.get("result", "")

    # コードブロック（```json ... ```）が含まれる場合は中身だけ取り出す
    m = re.search(r"```(?:json)?\s*(\{.*?\}|\[.*?\])\s*```", raw_text, re.DOTALL)
    json_str = m.group(1) if m else raw_text.strip()

    # {"items":[...]} 形式か [...] 形式かを判定
    if not (json_str.startswith("{") or json_str.startswith("[")):
        m2 = re.search(r"(\{.*\}|\[.*\])", json_str, re.DOTALL)
        if m2:
            json_str = m2.group(1)

    try:
        parsed = json.loads(json_str)
    except json.JSONDecodeError:
        # JSONが途中で切れている場合（項目数が多いPDFでレスポンスが長くなった場合など）、
        # 完結している品目オブジェクトだけを正規表現で抽出して部分復元を試みる
        parsed = _recover_partial_json(json_str, raw_text)

    # {"items": [...], "notes": "...", "大項目": "..."} または [...] どちらも受け付ける
    notes = ""
    daikomi = ""
    if isinstance(parsed, dict):
        items = parsed.get("items", [])
        notes = str(parsed.get("notes", "")).strip()
        daikomi = str(parsed.get("大項目", "")).strip()
    elif isinstance(parsed, list):
        items = parsed
    else:
        raise ClaudeExtractionError("Claude の応答が配列・オブジェクト形式ではありませんでした。")

    if not isinstance(items, list):
        raise ClaudeExtractionError("Claude の応答が配列形式ではありませんでした。")

    return items, notes, daikomi, result.get("total_cost_usd")


def extract_items_from_pdf(file_bytes: bytes) -> tuple[list[dict], str, str]:
    """PDFをClaudeに直接読み込ませて品目データを抽出する。

    Claude Code CLI の Read ツールを使ってPDFを読み取る。
    スキャンPDFは事前に向き（縦横・回転）を自動補正してから読ませる。
    APIが混雑している場合は時間がかかることがある（最大10分）。
    """
    try:
        from pdf_orient import ensure_upright_pdf
        file_bytes = ensure_upright_pdf(file_bytes)
    except Exception:
        pass  # 向き補正に失敗しても元PDFで続行
    with tempfile.TemporaryDirectory(prefix="quote_pdf_") as tmp_dir:
        tmp_path = Path(tmp_dir) / "quote.pdf"
        tmp_path.write_bytes(file_bytes)
        prompt = _CLAUDE_PROMPT.format(filename=tmp_path.name)
        cmd = [
            CLAUDE_BIN, "-p", prompt,
            "--output-format", "json",
            "--tools", "Read",
            "--add-dir", tmp_dir,
            "--dangerously-skip-permissions",
            "--model", "sonnet",
        ]
        result = _run_claude(cmd, timeout=CLAUDE_TIMEOUT_SEC, cwd=tmp_dir)

    raw_items, notes, daikomi, cost = _parse_claude_result(result)
    entries = _build_entries_from_claude_items(raw_items)
    if daikomi:
        entries.insert(0, {
            "種別": "大項目", "品名": daikomi,
            "規格": "", "備考": "", "単位": "", "数量": 0, "元単価": 0,
        })
    cost_str = f"（コスト: ${cost:.3f}）" if cost is not None else ""
    return entries, f"Claude AI解析{cost_str}", notes


def _build_entries_from_claude_items(raw_items: list[dict]) -> list[dict]:
    """Claudeが返した品目リストをアプリ内部の表現に変換する。"""
    entries: list[dict] = []
    prev_section: str | None = None

    # No.列の有無判定: 種別="大項目"の見出し行は除いて判定する
    has_no_column = any(
        str(raw.get("no", "")).strip()
        for raw in raw_items
        if str(raw.get("種別", "")).strip() != "大項目"
    )
    seen_nos: set[str] = set()

    for raw in raw_items:
        name = str(raw.get("品名", "")).strip()
        if not name:
            continue

        section = str(raw.get("種別", "")).strip()

        # ＜＞囲み見出し行（大項目センチネル）: 品目行とは別に大項目ヘッダーとして追加
        if section == "大項目":
            entries.append({
                "種別": "大項目", "品名": name,
                "規格": "", "備考": "", "単位": "", "数量": 0, "元単価": 0,
            })
            prev_section = None
            continue

        if name != "値引" and any(m in name for m in _SUMMARY_ROW_MARKERS):
            continue

        no = str(raw.get("no", "")).strip()

        # 種別（工事区分）に属する品目は常にサブ項目扱い（番号なし）。
        # 内部番号が外側番号と衝突しても seen_nos に追加せず、
        # 種別のない品目だけ seen_nos で重複判定する。
        if section:
            force_sub = True
        elif has_no_column:
            if no == "" or no in seen_nos:
                force_sub = True
            else:
                force_sub = False
                if no:
                    seen_nos.add(no)
        else:
            force_sub = False

        # セクション変化は force_sub に関係なく常に処理する。
        # こうしないと内部番号が外側番号と衝突した場合にヘッダーが遅れて挿入される。
        if section != prev_section:
            if prev_section is not None:
                entries.append({"種別": "小計区切り"})
            if section:
                entries.append({"種別": "工事区分", "品名": section, "no": ""})
            prev_section = section

        qty = to_number(raw.get("数量"))
        price = to_number(raw.get("単価"))
        amount = to_number(raw.get("金額"))

        eff_qty = qty if qty else 1
        if amount is not None:
            eff_price = amount / eff_qty if eff_qty else amount
        elif price is not None:
            eff_price = price
        else:
            continue

        entries.append({
            "種別": "品目",
            "no": no if not force_sub else "",
            "品名": name,
            "規格": str(raw.get("規格仕様", "")).strip(),
            "備考": str(raw.get("備考", "")).strip(),
            "単位": str(raw.get("単位", "")).strip(),
            "数量": eff_qty,
            "元単価": round(eff_price, 2),
            "force_sub": force_sub,
        })

    return entries


# ── Excel: openpyxl による品目抽出 ──────────────────────────────────

def extract_items_from_excel(file_bytes: bytes) -> list[dict]:
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    header_idx = col_name = col_qty = col_unit = col_spec = col_price = col_note = None
    for i, row in enumerate(rows[:15]):
        texts = [str(c).strip() if c is not None else "" for c in row]
        cn = find_col(texts, ITEM_NAME_KEYWORDS)
        cq = find_col(texts, QTY_KEYWORDS)
        cu = find_col(texts, UNIT_KEYWORDS)
        cs = find_col(texts, SPEC_KEYWORDS)
        cp = find_col(texts, PRICE_KEYWORDS)
        cnote = find_col(texts, ["備考", "note", "remarks"])
        if cn is not None and (cq is not None or cp is not None):
            header_idx, col_name, col_qty, col_unit, col_spec, col_price, col_note = (
                i, cn, cq, cu, cs, cp, cnote,
            )
            break

    if header_idx is None:
        return []

    def _cell_text(row, col):
        if col is None or col >= len(row) or row[col] is None:
            return ""
        return str(row[col]).strip()

    items = []
    for row in rows[header_idx + 1:]:
        if col_name >= len(row):
            continue
        name = row[col_name]
        if name is None or str(name).strip() == "":
            continue
        qty = to_number(row[col_qty]) if col_qty is not None and col_qty < len(row) else None
        price = to_number(row[col_price]) if col_price is not None and col_price < len(row) else None
        items.append({
            "種別": "品目",
            "品名": str(name).strip(),
            "規格": _cell_text(row, col_spec),
            "備考": _cell_text(row, col_note),
            "単位": _cell_text(row, col_unit),
            "数量": qty if qty is not None else 1,
            "元単価": price if price is not None else 0,
        })
    return items
