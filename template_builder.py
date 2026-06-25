from __future__ import annotations

import io
import math
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import range_boundaries
from openpyxl.worksheet.pagebreak import Break

import template_config as cfg

THIN = Side(style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="D9E2F3")   # 列見出し（青系）
SECTION_FILL = PatternFill("solid", fgColor="E2EFDA")  # 工事区分見出し（緑系）
DAIKOMI_FILL = PatternFill("solid", fgColor="1F3864")  # 大項目（ダークネイビー）
DAIKOMI_FONT = Font(bold=True, color="FFFFFFFF")        # 大項目の白字（ARGB）
CENTER = Alignment(horizontal="center", vertical="center")

# 品目データ行の列構成（結合する終了列。Noneは単一セル）
DATA_COL_SPECS = [
    (cfg.COL_ITEM_NAME, "C"), (cfg.COL_SPEC, "F"), (cfg.COL_QTY, None),
    (cfg.COL_UNIT, None), (cfg.COL_UNIT_PRICE, None), (cfg.COL_AMOUNT, "K"), (cfg.COL_NOTE, "N"),
]


def _merge(ws, cell_range: str):
    ws.merge_cells(cell_range)


def _set(ws, cell_range: str, value=None, font=None, fill=None, align=None, border=None, number_format=None):
    """セル（結合の場合は左上セル）に値・書式を設定するヘルパー。"""
    if ":" in cell_range:
        _merge(ws, cell_range)
        top_left = cell_range.split(":")[0]
    else:
        top_left = cell_range
    cell = ws[top_left]
    if value is not None:
        cell.value = value
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if align:
        cell.alignment = align
    if number_format:
        cell.number_format = number_format
    if border:
        min_col, min_row, max_col, max_row = range_boundaries(cell_range)
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                ws.cell(row=r, column=c).border = border
    return cell


NORMAL_FONT = Font(size=11, bold=False)
NORMAL_ALIGN = Alignment(horizontal="general", vertical="bottom")
NO_FILL = PatternFill(fill_type=None)


_DATA_NUMBER_FORMATS = {
    cfg.COL_UNIT_PRICE: "#,##0",
    cfg.COL_AMOUNT: "#,##0",
}


def _style_data_row(ws, row: int):
    """指定した1行を、品目データ行と同じ結合・罫線・フォントにする（流用元の太字や
    フォントサイズが残らないよう、フォント・配置・塗りつぶしも明示的にリセットする）。
    """
    for start_col, end_col in DATA_COL_SPECS:
        cell_range = f"{start_col}{row}:{end_col}{row}" if end_col else f"{start_col}{row}"
        number_format = _DATA_NUMBER_FORMATS.get(start_col)
        _set(ws, cell_range, font=NORMAL_FONT, fill=NO_FILL, align=NORMAL_ALIGN, border=BORDER,
             number_format=number_format)


def _clear_row_range(ws, row_start: int, row_end: int):
    """指定した行範囲の値・結合をすべて解除する（サマリー欄を移動する前の下準備）。"""
    for m in list(ws.merged_cells.ranges):
        if m.min_row >= row_start and m.max_row <= row_end:
            ws.unmerge_cells(str(m))
    for row in range(row_start, row_end + 1):
        for col in range(1, 15):
            ws.cell(row=row, column=col).value = None


def _build_summary_block(ws, subtotal_row: int, tax_row: int, total_row: int, tax_rate: float = 10):
    """小計・消費税・合計の3行ブロックを指定した行位置に作成する。"""
    header_font = Font(bold=True)

    _set(ws, f"A{subtotal_row}:{cfg.LABEL_COL_END_SUBTOTAL}{subtotal_row}", "小　計",
         font=header_font, align=Alignment(horizontal="center"), border=BORDER)
    _set(ws, f"{cfg.VALUE_COL_START}{subtotal_row}:{cfg.VALUE_COL_END}{subtotal_row}", 0,
         align=Alignment(horizontal="right"), border=BORDER, number_format="#,##0")

    _set(ws, f"A{tax_row}:{cfg.LABEL_COL_END_SUBTOTAL}{tax_row}", "消　費　税",
         font=header_font, align=Alignment(horizontal="center"), border=BORDER)
    _set(ws, f"{cfg.TAX_RATE_COL}{tax_row}", tax_rate, align=CENTER, border=BORDER)
    _set(ws, f"{cfg.TAX_RATE_UNIT_COL}{tax_row}", "％", align=CENTER, border=BORDER)
    _set(ws, f"{cfg.VALUE_COL_START}{tax_row}:{cfg.VALUE_COL_END}{tax_row}", 0,
         align=Alignment(horizontal="right"), border=BORDER, number_format="#,##0")

    total_row_end = total_row + 1
    _set(ws, f"A{total_row}:{cfg.LABEL_COL_END_TOTAL}{total_row_end}", "合　計",
         font=Font(bold=True, size=13), align=CENTER, border=BORDER)
    _set(ws, f"{cfg.VALUE_COL_START}{total_row}:{cfg.VALUE_COL_END}{total_row_end}", 0,
         font=Font(bold=True, size=13), align=Alignment(horizontal="right", vertical="center"),
         border=BORDER, number_format="#,##0")


def _build_sheet(wb, sheet_name: str):
    """見積書・請求書で共通のシートレイアウトを作る（タイトル・挨拶文・金額
    ラベルは sheet_name に応じて切り替える）。請求書シートのみ振込先欄を追加する。
    """
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)

    title_font = Font(size=20, bold=True)
    header_font = Font(bold=True)
    small_font = Font(size=9)
    bold_font = Font(bold=True)

    title_text = cfg.DOCUMENT_TITLES.get(sheet_name, sheet_name)
    greeting_text = cfg.DOCUMENT_GREETINGS.get(sheet_name, "")
    total_label = cfg.DOCUMENT_TOTAL_LABELS.get(sheet_name, "合計金額")

    # タイトル
    _set(ws, "A1:N2", title_text, font=title_font, align=Alignment(horizontal="center", vertical="center"))

    # 宛先
    _set(ws, f"{cfg.CELL_CLIENT}:D4", "（宛先を入力）", font=Font(size=14),
         align=Alignment(horizontal="left", vertical="bottom"))
    _set(ws, f"{cfg.CELL_CLIENT_SUFFIX}:E4", "様", font=Font(size=14), align=CENTER)
    _set(ws, f"{cfg.CELL_ISSUE_DATE}:N4", None, align=Alignment(horizontal="right"), number_format="yyyy年mm月dd日")

    # 許可番号など（任意）
    _set(ws, f"{cfg.CELL_ISSUER_LICENSE}:N6", "", font=small_font, align=Alignment(horizontal="right"))

    # 挨拶文
    _set(ws, cfg.CELL_GREETING, greeting_text)

    # 発行元情報
    _set(ws, f"{cfg.CELL_ISSUER_NAME}:N8", "", font=bold_font, align=Alignment(horizontal="right"))
    _set(ws, f"{cfg.CELL_ISSUER_ADDRESS}:N9", "", align=Alignment(horizontal="right", wrap_text=True))
    ws.row_dimensions[9].height = 28  # 折り返し2行分の高さを確保
    _set(ws, cfg.CELL_ISSUER_TEL, "", align=Alignment(horizontal="right"))
    _set(ws, cfg.CELL_ISSUER_FAX, "", align=Alignment(horizontal="right"))
    _set(ws, cfg.CELL_ISSUER_REG_NO, "", align=Alignment(horizontal="right"))

    # 金額サマリー（見積金額／請求金額）
    _set(ws, f"{cfg.CELL_TOTAL_LABEL}:D14", total_label, font=header_font, align=CENTER, border=BORDER)
    _set(ws, f"{cfg.CELL_TOTAL_DISPLAY}:J14", 0, font=Font(size=16, bold=True),
         align=Alignment(horizontal="center", vertical="center"), border=BORDER, number_format='#,##0"円（税込）"')
    # K13:K14 は検印欄との間のバッファ（空白）

    # 検印欄
    _set(ws, "L13:N13", "検　印", font=header_font, align=CENTER, border=BORDER)
    for col in ("L", "M", "N"):
        _set(ws, f"{col}14:{col}16", "", border=BORDER)

    # 品目テーブルのヘッダー
    headers = {
        f"{cfg.COL_ITEM_NAME}:C": "名称",
        f"{cfg.COL_SPEC}:F": "寸法・規格",
        cfg.COL_QTY: "数量",
        cfg.COL_UNIT: "単位",
        cfg.COL_UNIT_PRICE: "単価",
        f"{cfg.COL_AMOUNT}:K": "金額",
        f"{cfg.COL_NOTE}:N": "備考",
    }
    header_row = cfg.COLUMN_HEADER_ROW
    for col_spec, label in headers.items():
        if ":" in col_spec:
            start_col, end_col = col_spec.split(":")
            cell_range = f"{start_col}{header_row}:{end_col}{header_row}"
        else:
            cell_range = f"{col_spec}{header_row}"
        _set(ws, cell_range, label, font=header_font, fill=HEADER_FILL, align=CENTER, border=BORDER)

    # 品目データ行（罫線・結合のみ用意）
    for r in range(cfg.DATA_START_ROW, cfg.DATA_START_ROW + cfg.MAX_DATA_ROWS):
        _style_data_row(ws, r)

    # 小計・消費税・合計
    _build_summary_block(ws, cfg.SUBTOTAL_ROW, cfg.TAX_ROW, cfg.TOTAL_ROW)

    # 振込先（請求書のみ）
    if sheet_name == cfg.INVOICE_SHEET_NAME:
        _build_bank_row(ws, cfg.TOTAL_ROW + cfg.BANK_ROW_OFFSET)

    widths = {
        "A": 12, "B": 6, "C": 6, "D": 6, "E": 5, "F": 6,
        "G": 5, "H": 4, "I": 8, "J": 7, "K": 5, "L": 6, "M": 5, "N": 6,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    _setup_a4_print(ws)
    return ws


def _setup_a4_print(ws):
    """A4縦・余白・1ページ幅フィット・列見出し繰り返しの印刷設定。"""
    ws.page_setup.paperSize = 9          # A4
    ws.page_setup.orientation = "portrait"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1         # 横は必ず1ページ幅に収める
    ws.page_setup.fitToHeight = 0        # 縦は複数ページ可
    ws.page_margins.left = 0.71          # 約18mm
    ws.page_margins.right = 0.71
    ws.page_margins.top = 0.79           # 約20mm
    ws.page_margins.bottom = 0.79
    ws.page_margins.header = 0.31
    ws.page_margins.footer = 0.31
    # 2ページ目以降にも列見出し行を繰り返す
    ws.print_title_rows = f"{cfg.COLUMN_HEADER_ROW}:{cfg.COLUMN_HEADER_ROW}"


def _build_notes_block(ws, start_row: int, notes: str):
    """備考テキストブロックを合計欄の下に追記する。"""
    if not notes.strip():
        return

    r = start_row + 1  # 1行空ける

    # 「備考」ヘッダー行
    _set(ws, f"A{r}:N{r}", "備　考",
         font=Font(bold=True),
         fill=HEADER_FILL,
         align=Alignment(horizontal="left", vertical="center"),
         border=BORDER)
    ws.row_dimensions[r].height = 18
    r += 1

    # テキスト本文：改行ごとに1行として出力
    for line_text in notes.strip().split("\n"):
        _set(ws, f"A{r}:N{r}", line_text,
             align=Alignment(horizontal="left", vertical="center", wrap_text=True),
             border=BORDER)
        ws.row_dimensions[r].height = 15
        r += 1


def _build_bank_row(ws, row: int):
    """振込先（銀行名・支店・口座番号など）の行を作る。"""
    header_font = Font(bold=True)
    _set(ws, f"{cfg.CELL_BANK_LABEL_COL}{row}", "振込先", font=header_font,
         align=Alignment(horizontal="left", vertical="center"))
    _set(ws, f"{cfg.CELL_BANK_INFO_COL_START}{row}:{cfg.CELL_BANK_INFO_COL_END}{row}", "",
         align=Alignment(horizontal="left", vertical="center"))


def ensure_template_exists(path: Path | None = None) -> Path:
    """起動のたびにテンプレートを再生成する（列幅・印刷設定を常に最新に保つ）。"""
    path = Path(path) if path else Path(cfg.TEMPLATE_PATH)
    wb = Workbook()
    wb.remove(wb.active)
    _build_sheet(wb, cfg.SHEET_NAME)
    _build_sheet(wb, cfg.INVOICE_SHEET_NAME)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def _build_render_lines(items: list[dict], show_section_subtotals: bool = False) -> tuple[list[dict], int]:
    """items から実際にシートへ書き出す行のリストを組み立てる。

    - 種別="大項目"  → 最上位ヘッダー行（kind="daikomi"）
    - 種別="工事区分" → 番号付き見出し行（kind="section"）
    - 種別="小計区切り" → セクション境界（行は生成しない）
    - show_section_subtotals=True のときのみ【小計】行を挿入する
    各品目行には is_sub_item フラグを付与し _fill_sheet 側でインデントを制御する。
    """
    has_markers = any(e.get("種別") in ("工事区分", "小計区切り", "大項目") for e in items)

    def _item_line(entry: dict) -> tuple[dict, int]:
        qty = entry.get("数量", 0) or 0
        price = math.floor(entry.get("上乗せ後単価", 0) or 0)
        amount = math.floor(qty * price)
        line = {
            "kind": "item",
            "no": str(entry.get("no", "")).strip(),
            "品名": entry.get("品名", ""), "規格": entry.get("規格", ""),
            "単位": entry.get("単位", ""), "数量": qty, "単価": price, "金額": amount,
            "備考": entry.get("備考", ""), "is_sub_item": False,
        }
        return line, amount

    lines: list[dict] = []
    subtotal = 0

    if not has_markers:
        for entry in items:
            line, amount = _item_line(entry)
            subtotal += amount
            lines.append(line)
        return lines, subtotal

    section_amount = 0
    section_has_items = False
    in_section_flag = False

    def close_section():
        nonlocal section_amount, section_has_items, in_section_flag
        if section_has_items and show_section_subtotals and section_amount > 0:
            lines.append({"kind": "section_subtotal", "amount": section_amount})
        section_amount = 0
        section_has_items = False
        in_section_flag = False

    for entry in items:
        kind = entry.get("種別")
        if kind == "大項目":
            close_section()
            lines.append({"kind": "daikomi", "label": entry.get("品名", "")})
        elif kind == "工事区分":
            close_section()
            in_section_flag = True
            lines.append({"kind": "section", "label": entry.get("品名", ""), "no": str(entry.get("no", "")).strip()})
        elif kind == "小計区切り":
            close_section()
        else:
            line, amount = _item_line(entry)
            line["is_sub_item"] = True if entry.get("force_sub") else in_section_flag
            subtotal += amount
            section_amount += amount
            section_has_items = True
            lines.append(line)
    close_section()
    return lines, subtotal


def _fill_sheet(ws, items: list[dict], header_info: dict, issuer_info: dict,
                 tax_rate: float, is_invoice: bool, bank_info: str = "", notes: str = "",
                 show_section_subtotals: bool = False):
    """1枚のシート（見積書 または 請求書）にデータを流し込む共通処理。"""
    if header_info.get("client"):
        ws[cfg.CELL_CLIENT] = header_info["client"]
    if header_info.get("issue_date"):
        ws[cfg.CELL_ISSUE_DATE] = header_info["issue_date"]

    if issuer_info.get("license"):
        ws[cfg.CELL_ISSUER_LICENSE] = issuer_info["license"]
    if issuer_info.get("name"):
        ws[cfg.CELL_ISSUER_NAME] = issuer_info["name"]
    if issuer_info.get("address"):
        ws[cfg.CELL_ISSUER_ADDRESS] = issuer_info["address"]
    if issuer_info.get("tel"):
        ws[cfg.CELL_ISSUER_TEL] = f"TEL　{issuer_info['tel']}"
    if issuer_info.get("fax"):
        ws[cfg.CELL_ISSUER_FAX] = f"FAX　{issuer_info['fax']}"
    if issuer_info.get("registration_no"):
        ws[cfg.CELL_ISSUER_REG_NO] = f"登録番号：{issuer_info['registration_no']}"

    render_lines, subtotal = _build_render_lines(items, show_section_subtotals=show_section_subtotals)

    subtotal_row, tax_row, total_row = cfg.SUBTOTAL_ROW, cfg.TAX_ROW, cfg.TOTAL_ROW
    bank_row = cfg.TOTAL_ROW + cfg.BANK_ROW_OFFSET
    overflow = max(0, len(render_lines) - cfg.MAX_DATA_ROWS)

    if overflow > 0:
        # 元の小計〜合計（請求書なら振込先まで）ブロックを解除する
        old_block_end = bank_row if is_invoice else cfg.TOTAL_ROW + 1
        _clear_row_range(ws, cfg.SUBTOTAL_ROW, old_block_end)

        subtotal_row = cfg.SUBTOTAL_ROW + overflow
        tax_row = cfg.TAX_ROW + overflow
        total_row = cfg.TOTAL_ROW + overflow
        bank_row = total_row + cfg.BANK_ROW_OFFSET

        # あらかじめ罫線を用意した品目行より下、新しい小計欄より上の行は
        # すべて通常のデータ行としてスタイルを適用する（途中に未装飾の
        # 空白行を残さないよう、ギャップなく連続した範囲で処理する）
        for row in range(cfg.DATA_START_ROW + cfg.MAX_DATA_ROWS, subtotal_row):
            _style_data_row(ws, row)

        _build_summary_block(ws, subtotal_row, tax_row, total_row, tax_rate=tax_rate)
        if is_invoice:
            _build_bank_row(ws, bank_row)

    # A4印刷設定を出力ファイルにも確実に適用
    _setup_a4_print(ws)

    # 縦は複数ページ可（fitToHeight=1にすると品目が少ない場合でも1ページに圧縮され文字が小さくなるため0固定）
    ws.page_setup.fitToHeight = 0

    # ページ区切り: 大項目/工事区分の見出しがページ末尾に孤立しないよう直前で改ページ。
    # 強制改ページは入れず Excelの自動改ページに任せることで
    # 1ページ目をできるだけ詰めて余白を最小化する。
    # A4縦・ヘッダー17行差引後の推定データ行数: 1p目≒30行、2p目以降≒45行
    _PAGE1_DATA_ROWS = 30
    _PAGEX_DATA_ROWS = 45
    page_capacity = _PAGE1_DATA_ROWS
    rows_on_page = 0

    bold = Font(bold=True)
    item_counter = 0
    for i, line in enumerate(render_lines):
        row = cfg.DATA_START_ROW + i

        # 見出し行がページ残量 20% 未満の位置に来る場合のみ直前に改ページ
        if line["kind"] in ("daikomi", "section") and rows_on_page >= int(page_capacity * 0.8):
            ws.row_breaks.append(Break(id=row - 1))
            rows_on_page = 0
            page_capacity = _PAGEX_DATA_ROWS

        rows_on_page += 1
        # ページ境界を越えた場合は次ページの容量にリセット（改ページは挿入しない）
        if rows_on_page >= page_capacity:
            rows_on_page = 0
            page_capacity = _PAGEX_DATA_ROWS

        if line["kind"] == "daikomi":
            # 大項目ヘッダー：ダークネイビー背景・白字・番号なし。配下品目の番号はリセット
            item_counter = 0
            for col_num in range(1, 15):
                ws.cell(row=row, column=col_num).fill = DAIKOMI_FILL
            cell = ws[f"{cfg.COL_ITEM_NAME}{row}"]
            cell.value = line["label"]
            cell.font = DAIKOMI_FONT
        elif line["kind"] == "section":
            # 工事区分見出し：元PDFの番号があればそれを使い、なければ通し番号
            cell = ws[f"{cfg.COL_ITEM_NAME}{row}"]
            no = line.get("no", "")
            if no:
                cell.value = f"{no}. {line['label']}" if line["label"] else ""
                try:
                    item_counter = int(no)
                except (ValueError, TypeError):
                    pass
            else:
                item_counter += 1
                cell.value = f"{item_counter}. {line['label']}" if line["label"] else ""
        elif line["kind"] == "section_subtotal":
            name_cell = ws[f"{cfg.COL_ITEM_NAME}{row}"]
            name_cell.value = "【小計】"
            name_cell.font = bold
            amount_cell = ws[f"{cfg.COL_AMOUNT}{row}"]
            amount_cell.value = line["amount"]
            amount_cell.font = bold
        else:  # item
            if line.get("is_sub_item"):
                # 工事区分配下のサブ項目：番号なし・インデント
                item_name = f"　{line['品名']}" if line["品名"] else ""
            else:
                # 元PDFの番号があればそれを使い、なければ通し番号
                # item_counter を明示番号と常に同期させることで、
                # 番号なし見出し（工事区分）が次の連番を正しく引き継げる
                no = line.get("no", "")
                if no:
                    item_name = f"{no}. {line['品名']}" if line["品名"] else ""
                    try:
                        item_counter = int(no)
                    except (ValueError, TypeError):
                        pass
                else:
                    item_counter += 1
                    item_name = f"{item_counter}. {line['品名']}" if line["品名"] else ""
            ws[f"{cfg.COL_ITEM_NAME}{row}"] = item_name
            ws[f"{cfg.COL_SPEC}{row}"] = line["規格"]
            ws[f"{cfg.COL_UNIT}{row}"] = line["単位"]
            ws[f"{cfg.COL_QTY}{row}"] = line["数量"]
            ws[f"{cfg.COL_UNIT_PRICE}{row}"] = line["単価"]
            ws[f"{cfg.COL_AMOUNT}{row}"] = line["金額"]
            ws[f"{cfg.COL_NOTE}{row}"] = line["備考"]

    tax_amount = math.floor(subtotal * tax_rate / 100)
    grand_total = subtotal + tax_amount

    ws[f"{cfg.VALUE_COL_START}{subtotal_row}"] = subtotal
    ws[f"{cfg.TAX_RATE_COL}{tax_row}"] = tax_rate
    ws[f"{cfg.VALUE_COL_START}{tax_row}"] = tax_amount
    ws[f"{cfg.VALUE_COL_START}{total_row}"] = grand_total
    ws[cfg.CELL_TOTAL_DISPLAY] = grand_total

    if is_invoice and bank_info:
        ws[f"{cfg.CELL_BANK_INFO_COL_START}{bank_row}"] = bank_info

    # 備考ブロック
    notes_start = bank_row if is_invoice else total_row + 1
    _build_notes_block(ws, notes_start, notes)


def fill_template(items: list[dict], header_info: dict, issuer_info: dict,
                   tax_rate: float, template_path: Path | None = None,
                   document_type: str | None = None, bank_info: str = "") -> bytes:
    """items: [{"種別","品名","規格","数量","単位","上乗せ後単価"}]（税別）を
    template の1シートに流し込み、bytes を返す（見積書のみ／請求書のみが
    必要な場合に使う。両方まとめて1冊にしたい場合は fill_combined_document
    を使う）。
    """
    document_type = document_type or cfg.SHEET_NAME
    is_invoice = document_type == cfg.INVOICE_SHEET_NAME

    template_path = Path(template_path) if template_path else Path(cfg.TEMPLATE_PATH)
    wb = load_workbook(template_path)
    ws = wb[document_type] if document_type in wb.sheetnames else wb.active

    _fill_sheet(ws, items, header_info, issuer_info, tax_rate, is_invoice, bank_info)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def fill_combined_document(items: list[dict], header_info: dict, issuer_info: dict,
                            tax_rate: float, template_path: Path | None = None,
                            bank_info: str = "", notes: str = "",
                            show_section_subtotals: bool = False,
                            invoice_header_info: dict | None = None) -> bytes:
    """同じ品目・税率から「見積書」「請求書」両方のシートを1冊のExcelに
    まとめて出力する（1回のダウンロードで両方そろう）。
    invoice_header_info を渡すと請求書に別の発行日・宛先を使える。
    """
    template_path = Path(template_path) if template_path else Path(cfg.TEMPLATE_PATH)
    wb = load_workbook(template_path)
    inv_hdr = invoice_header_info if invoice_header_info is not None else header_info

    quote_ws = wb[cfg.SHEET_NAME] if cfg.SHEET_NAME in wb.sheetnames else wb.active
    _fill_sheet(quote_ws, items, header_info, issuer_info, tax_rate, is_invoice=False,
                notes=notes, show_section_subtotals=show_section_subtotals)

    if cfg.INVOICE_SHEET_NAME in wb.sheetnames:
        invoice_ws = wb[cfg.INVOICE_SHEET_NAME]
        _fill_sheet(invoice_ws, items, inv_hdr, issuer_info, tax_rate, is_invoice=True,
                    bank_info=bank_info, notes=notes, show_section_subtotals=show_section_subtotals)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
